# coding=utf-8
"""
Created on 2018-01-02

@Filename: users_check
@Author: Gui


"""
import logging
from openpyxl import Workbook
from epMongo import EPMongo


def generate_pipeline(user_id):
    pipeline = [
        {
            "$match": {
                "_id": user_id
            }
        },

        {
            "$lookup": {
                "from": "userpositions",
                "localField": "position",
                "foreignField": "_id",
                "as": "position"
            }
        },

        {
            "$unwind": {
                "path": "$position",
                "preserveNullAndEmptyArrays": True
            }
        },

        {
            "$lookup": {
                "from": "usergroups",
                "localField": "position.group",
                "foreignField": "_id",
                "as": "position.group"
            }
        },

        {
            "$unwind": {
                "path": "$position.group",
                "preserveNullAndEmptyArrays": True
            }
        },

        {
            "$lookup": {
                "from": "accounts",
                "localField": "position.group.account",
                "foreignField": "_id",
                "as": "position.group.account"
            }
        },

        {
            "$unwind": {
                "path": "$position.group.account",
                "preserveNullAndEmptyArrays": True
            }
        },

        {
            "$lookup": {
                "from": "organizations",
                "localField": "position.group.account.organization",
                "foreignField": "_id",
                "as": "position.group.account.organization"
            }
        },

        {
            "$unwind": {
                "path": "$position.group.account.organization",
                "preserveNullAndEmptyArrays": True
            }
        },

        {
            "$project": {
                "userId": "$_id",
                "username": 1,
                "dateDelete": 1,
                "isSuperAdmin": 1,
                "status": 1,
                "group_id": "$position.group._id",
                "group_name": "$position.group.name",
                "group_default": "$position.group.default",
                "group_dateDelete": "$position.group.dateDelete",
                "account_id": "$position.group.account._id",
                "account_name": "$position.group.account.name",
                "account_types": "$position.group.account.types",
                "organization_id": "$position.group.account.organization._id",
                "organization_name": "$position.group.account.organization.name",
                "organization_dateDelete": "$position.group.account.organization.dateDelete",
            }
        },

    ]
    return pipeline


def dict_has_key(dict, key):
    if key in dict.keys():
        value = dict[key]
        if isinstance(value, (list,)):
            value = value[0]
        return value
    else:
        return '--'


if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG)
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = 'No.'
    ws.cell(row=1, column=2).value = '_id'
    ws.cell(row=1, column=3).value = 'username'
    ws.cell(row=1, column=4).value = 'dateDelete'
    ws.cell(row=1, column=5).value = 'isSuperAdmin'
    ws.cell(row=1, column=6).value = 'status'
    ws.cell(row=1, column=7).value = 'group_id'
    ws.cell(row=1, column=8).value = 'group_name'
    ws.cell(row=1, column=9).value = 'group_default'
    ws.cell(row=1, column=10).value = 'group_dateDelete'
    ws.cell(row=1, column=11).value = 'account_id'
    ws.cell(row=1, column=12).value = 'account_name'
    ws.cell(row=1, column=13).value = 'account_types'
    ws.cell(row=1, column=14).value = 'organization_id'
    ws.cell(row=1, column=15).value = 'organization_name'
    ws.cell(row=1, column=16).value = 'organization_dateDelete'

    # epmongo = EPMongo()
    epmongo = EPMongo(
        uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
        db_name='epdb-prod')
    users = epmongo.get_docs_by_query('users', query=None, projection={'_id': 1})
    j = 1
    logging.info('Total %d users.' % users.count())
    for i, each in enumerate(users, 1):
        user_id = each['_id']
        pipeline = generate_pipeline(user_id)
        logging.info("{} {}".format(i, str(user_id)))
        for result in (epmongo.aggregate('users', pipeline)):
            logging.info(result)
            ws.cell(row=j + 1, column=1).value = i
            ws.cell(row=j + 1, column=2).value = str(dict_has_key(result, '_id'))
            ws.cell(row=j + 1, column=3).value = dict_has_key(result, 'username')
            ws.cell(row=j + 1, column=4).value = dict_has_key(result, 'dateDelete')
            ws.cell(row=j + 1, column=5).value = dict_has_key(result, 'isSuperAdmin')
            ws.cell(row=j + 1, column=6).value = dict_has_key(result, 'status')
            ws.cell(row=j + 1, column=7).value = str(dict_has_key(result, 'group_id'))
            ws.cell(row=j + 1, column=8).value = dict_has_key(result, 'group_name')
            ws.cell(row=j + 1, column=9).value = dict_has_key(result, 'group_default')
            ws.cell(row=j + 1, column=10).value = dict_has_key(result, 'group_dateDelete')
            ws.cell(row=j + 1, column=11).value = str(dict_has_key(result, 'account_id'))
            ws.cell(row=j + 1, column=12).value = dict_has_key(result, 'account_name')
            ws.cell(row=j + 1, column=13).value = dict_has_key(result, 'account_types')
            ws.cell(row=j + 1, column=14).value = str(dict_has_key(result, 'organization_id'))
            ws.cell(row=j + 1, column=15).value = dict_has_key(result, 'organization_name')
            ws.cell(row=j + 1, column=16).value = dict_has_key(result, 'organization_dateDelete')
            j = j + 1
    wb.save('%s.xlsx' % users.count())

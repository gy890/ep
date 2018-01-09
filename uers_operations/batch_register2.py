# coding=utf-8
"""
Created on 2017-12-12

@Filename: parse_excel
@Author: Gui


"""
import json
import time
from openpyxl import load_workbook
import requests


def parse_one(value):
    (a, b) = value.split('\n')
    (a, b) = (a.strip(), b.strip())
    return a, b


def enroll(port, organization, username):
    # url = 'http://dev.e-ports.com/api/'
    url = 'https://portal.e-ports.com/api/'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
        'Content-Type': 'application/json'}
    payload = {
        "action": "enroll",
        "request": {
            "company": organization,
            "emergencyEmail": username,
            "password": 'password',
            "port": port,
            "type": 'Agent'
        }
    }
    r = requests.post(url, data=json.dumps(payload), headers=headers)
    return r.status_code


def get_port_organization(port, user1, user2, user3, user4, user5, user6, user7, user8):
    port_list = []
    if user1 is not None:
        port_list.append((port, parse_one(user1)[0], parse_one(user1)[1]))
    if user2 is not None:
        port_list.append((port, parse_one(user2)[0], parse_one(user2)[1]))
    if user3 is not None:
        port_list.append((port, parse_one(user3)[0], parse_one(user3)[1]))
    if user4 is not None:
        port_list.append((port, parse_one(user4)[0], parse_one(user4)[1]))
    if user5 is not None:
        port_list.append((port, parse_one(user5)[0], parse_one(user5)[1]))
    if user6 is not None:
        port_list.append((port, parse_one(user6)[0], parse_one(user6)[1]))
    if user7 is not None:
        port_list.append((port, parse_one(user7)[0], parse_one(user7)[1]))
    if user8 is not None:
        port_list.append((port, parse_one(user8)[0], parse_one(user8)[1]))
    return port_list


if __name__ == '__main__':
    wb = load_workbook('第三批国内代理注册.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    organizations = []
    for i in range(3, 83):
        port = ws.cell(row=i, column=3).value
        user1 = ws.cell(row=i, column=5).value
        user2 = ws.cell(row=i, column=6).value
        user3 = ws.cell(row=i, column=7).value
        user4 = ws.cell(row=i, column=8).value
        user5 = ws.cell(row=i, column=9).value
        user6 = ws.cell(row=i, column=10).value
        user7 = ws.cell(row=i, column=11).value
        user8 = ws.cell(row=i, column=12).value
        organizations.extend(get_port_organization(port, user1, user2, user3, user4, user5, user6, user7, user8))
    for i, (port, organization, username) in enumerate(organizations, 1):
        # print(i, port, '-------', organization, '-------', username)
        print(username)
        # print(enroll(port, organization, username))

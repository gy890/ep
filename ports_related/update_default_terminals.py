# coding=utf-8
"""
Created on 2018-02-05

@Filename: update_default_terminals
@Author: Gui


"""
import logging
from bson.objectid import ObjectId
from openpyxl import load_workbook
from epMongo import EPMongo


def main():
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s %(message)s')
    epmongo = EPMongo()
    # epmongo = EPMongo(
    #     uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
    #     db_name='epdb-prod')

    wb = load_workbook('defaultTerminal.xlsx')
    ws = wb.active
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        default_terminal = {}
        port_id = ws.cell(row=row, column=3).value
        stdb = ws.cell(row=row, column=6).value
        stot = ws.cell(row=row, column=8).value
        if stdb:
            default_terminal.setdefault('STDB', ObjectId(stdb))
        if stot:
            default_terminal.setdefault('STOT', ObjectId(stot))
        if default_terminal:
            logging.info('row={} {}:{}'.format(row, port_id, default_terminal))
            logging.info(
                epmongo.update_one("ports", {'_id': ObjectId(port_id)},
                                   {"$set": {'defaultTerminal': default_terminal}}))


if __name__ == '__main__':
    main()

# coding=utf-8
"""
Created on 2018-02-05

@Filename: default_terminals
@Author: Gui


"""
import logging
from openpyxl import load_workbook


def get_ports_terminals():
    wb = load_workbook('ports(102pro).xlsx')
    ws = wb.active
    row = 2
    max_row = ws.max_row
    ports = {}
    for row in range(2, max_row + 1):
        port_id = ws.cell(row=row, column=2).value
        terminals_str = ws.cell(row=row, column=7).value
        terminals = {}
        for each in terminals_str.split('\n', ):
            terminal_id, name = each.split(',', maxsplit=1)
            terminals.setdefault(terminal_id, name)
        ports.setdefault(port_id, terminals)
    return ports


def set_default_terminals(ports):
    wb = load_workbook('defaultTerminal.xlsx')
    ws = wb.active
    row = 2
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        port_id = ws.cell(row=row, column=3).value
        stdb = ws.cell(row=row, column=5).value
        stot = ws.cell(row=row, column=7).value
        terminals = ports.get(port_id)
        if stdb:
            for k, v in terminals.items():
                if v.upper() == stdb.upper:
                    ws.cell(row=row, column=6).value = k
                    break
        if stot:
            for k, v in terminals.items():
                if v.upper() == stot.upper():
                    ws.cell(row=row, column=8).value = k
                    break
    wb.save('defaultTerminal.xlsx')


if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG,
                        format='%(asctime)s %(levelname)s %(message)s')
    ports = get_ports_terminals()
    set_default_terminals(ports)

# coding=utf-8
"""
Created on 2018-01-17

@Filename: update_geo
@Author: Gui


"""
import logging
from bson.objectid import ObjectId
from openpyxl import load_workbook
from epMongo import EPMongo

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s %(message)s')
    # epmongo = EPMongo()
    epmongo = EPMongo(
        uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
        db_name='epdb-prod')
    wb = load_workbook('ports.xlsx')
    ws = wb.active
    j = 1
    for i in range(1, ws.max_row + 1):
        port_id = ObjectId(ws.cell(row=i + 1, column=2).value)
        country = ws.cell(row=i + 1, column=5).value
        geo = ws.cell(row=i + 1, column=6).value
        if country == 'China':
            logging.debug((j, port_id, country, geo))
            j = j + 1
            logging.info(epmongo.update_one("ports", {'_id': port_id}, {"$set": {'geo': geo}}))

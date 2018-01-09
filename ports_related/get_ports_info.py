# coding=utf-8
"""
Created on 2018-01-03

@Filename: get_ports_info
@Author: Gui


"""
import logging
from openpyxl import Workbook, load_workbook
from epMongo import EPMongo


class PORTS(object):
    def __init__(self):
        # self._epmongo = EPMongo()
        self._epmongo = EPMongo(
            uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
            db_name='epdb-prod')

    def get_ports(self):
        self.ports = self._epmongo.get_docs_by_query('ports', {'dateDelete': {'$exists': False}},
                                                     {'name': 1, 'code': 1, 'country': 1, 'terminals': 1,
                                                      'shipyards': 1})
        return self.ports

    def _get_country(self, country_id):
        doc = self._epmongo.get_doc_by_query('countries', {'_id': country_id}, {'name': 1})
        return doc['name']

    def _get_terminal(self, terminal_id):
        doc = self._epmongo.get_doc_by_query('terminals', {'_id': terminal_id}, {'name': 1, '_id': 1})
        return doc

    def _get_shipyard(self, shipyard_id):
        doc = self._epmongo.get_doc_by_query('shipyards', {'_id': shipyard_id}, {'name': 1, '_id': 1})
        return doc

    def parse_one(self, port_doc):
        port_dict = {}
        _id = port_doc.get('_id')
        port_dict['_id'] = _id
        name = port_doc.get('name')
        port_dict['name'] = name
        code = port_doc.get('code')
        port_dict['code'] = code
        country = self._get_country(port_doc.get('country'))
        port_dict['country'] = country

        terminals = port_doc.get('terminals')
        if terminals is not None and len(terminals) != 0:
            terminals_list = []
            for terminal_id in terminals:
                terminal_doc = self._get_terminal(terminal_id)
                terminals_list.append(terminal_doc)
            port_dict['terminals'] = terminals_list
        shipyards = port_doc.get('shipyards')
        if shipyards is not None and len(shipyards) != 0:
            shipyards_list = []
            for shipyard_id in shipyards:
                shipyard_doc = self._get_shipyard(shipyard_id)
                shipyards_list.append(shipyard_doc)
            port_dict['shipyards'] = shipyards_list
        return port_dict


def convert_to_string(list):
    str_list = []
    for item in list:
        _id = str(item.get('_id'))
        name = item.get('name')
        id_name_string = ','.join((_id, name))
        str_list.append(id_name_string)
    return str_list


if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG)
    ports = PORTS()
    logging.info((type(ports.get_ports()), ports.get_ports()))

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = 'No.'
    ws.cell(row=1, column=2).value = '_id'
    ws.cell(row=1, column=3).value = 'name'
    ws.cell(row=1, column=4).value = 'code'
    ws.cell(row=1, column=5).value = 'country'
    ws.cell(row=1, column=6).value = 'terminals_number'
    ws.cell(row=1, column=7).value = 'terminals'
    ws.cell(row=1, column=8).value = 'shipyards_number'
    ws.cell(row=1, column=9).value = 'shipyards'

    i = 1
    for each in ports.get_ports():
        port_dict = ports.parse_one(each)
        logging.info((i, port_dict))
        ws.cell(row=i + 1, column=1).value = i
        ws.cell(row=i + 1, column=2).value = str(port_dict.get('_id'))
        ws.cell(row=i + 1, column=3).value = port_dict.get('name')
        ws.cell(row=i + 1, column=4).value = port_dict.get('code')
        ws.cell(row=i + 1, column=5).value = port_dict.get('country')

        terminals_list = port_dict.get('terminals')
        if terminals_list is not None:
            ws.cell(row=i + 1, column=6).value = len(terminals_list)
            ws.cell(row=i + 1, column=7).value = '\n'.join(convert_to_string(terminals_list))

        shipyards_list = port_dict.get('shipyards')
        if shipyards_list is not None:
            ws.cell(row=i + 1, column=8).value = len(shipyards_list)
            ws.cell(row=i + 1, column=9).value = '\n'.join(convert_to_string(shipyards_list))

        i = i + 1
    wb.save('ports.xlsx')

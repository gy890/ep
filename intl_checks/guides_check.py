# coding=utf-8
"""
Created on 2018-01-26

@Filename: guides_check
@Author: Gui


"""
import logging
import collections
import timeit
from openpyxl import Workbook


@timeit.timeit
def main():
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s: %(message)s')
    with open('guides-config.js') as f:
        lines = [l.strip() for l in f.readlines() if l.find('module.exports') == -1 and l.find('default') == -1]

    guides = []
    guide = collections.OrderedDict()
    for line in lines:
        if line.find('.') != -1:
            k = line.split("'")[1].strip()
            guide.setdefault('key', k)
            print(k, end='')

        if line.find('content') != -1:
            content = line.split("'")[1].strip()
            guide.setdefault('content', content)
            print(('\t{}'.format(content)), end='')

        if line.find('index') != -1:
            index = (line.split(":")[1]).strip(',').strip()
            guide.setdefault('index', index)
            print(('\t{}'.format(index)), end='')

        if line.find('position') != -1:
            position = line.split('"')[1].strip()
            guide.setdefault('position', position)
            print(('\t{}'.format(position)), end='')

        if line.find('covered') != -1:
            covered = (line.split(":")[1]).strip(',').strip()
            guide.setdefault('covered', covered)
            print(('\t{}'.format(covered)), end='')

        if line.find('circle') != -1:
            circle = (line.split(":")[1]).strip(',').strip()
            guide.setdefault('circle', circle)
            print(('\t{}'.format(circle)), end='')

        if line.find('},') != -1:
            a = guide.copy()
            guides.append(a)
            guide.clear()
            print()

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = 'No.'
    ws.cell(row=1, column=2).value = 'ref'
    ws.cell(row=1, column=3).value = 'intl'
    ws.cell(row=1, column=4).value = 'index'
    ws.cell(row=1, column=5).value = 'position'
    ws.cell(row=1, column=6).value = 'circle'
    ws.cell(row=1, column=7).value = 'coverd'
    i = 2
    for each in guides:
        print(each)
        ws.cell(row=i, column=1).value = i-1
        ws.cell(row=i, column=2).value = each.get('key')
        ws.cell(row=i, column=3).value = each.get('content')
        ws.cell(row=i, column=4).value = each.get('index')
        ws.cell(row=i, column=5).value = each.get('position')
        ws.cell(row=i, column=6).value = each.get('circle')
        ws.cell(row=i, column=7).value = each.get('circle')
        i = i + 1
    wb.save('guides.xlsx')

if __name__ == '__main__':
    main()

# coding=utf-8
"""
Created on 2018-04-08

@Filename: get_ordertypes
@Author: Gui


"""
from epMongo import EPMongo
from openpyxl import Workbook
import pymongo


def get_ordertypes():
    ep = EPMongo()
    # ep = EPMongo(
    #     uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
    #     db_name='epdb-prod')
    docs = ep.get_docs_by_query('ordertypes', query=None,
                                projection={'name': 1, 'code': 1, 'defaults': 1})
    docs.sort([('code', pymongo.ASCENDING)])
    products = ep.get_docs_by_query('products', query=None,
                                    projection={'name': 1, 'code': 1})
    products = {product.get('_id', ''): (product.get('name', ''), product.get('code', '')) for product in products}
    ordertypes = []
    for doc in docs:
        name = doc.get('name', '')
        code = doc.get('code', '')
        defaults = doc.get('defaults', '')
        default_products = []
        if defaults:
            for default in defaults:
                product_name, product_code = products.get(default, '')
                default_products.append((product_name, product_code))
        ordertypes.append((name, code, default_products))
    return ordertypes


if __name__ == '__main__':
    ordertypes = get_ordertypes()
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = 'No.'
    ws.cell(row=1, column=2).value = 'name'
    ws.cell(row=1, column=3).value = 'code'
    ws.cell(row=1, column=4).value = 'code'
    ws.cell(row=1, column=5).value = 'product_name'
    ws.cell(row=1, column=6).value = 'product_code'
    n = 1
    for i, ordertype in enumerate(ordertypes, 1):
        merge_start = n + 1
        name, code, default_products = ordertype
        ws.cell(row=n + 1, column=1).value = i
        ws.cell(row=n + 1, column=2).value = name
        ws.cell(row=n + 1, column=3).value = code

        if default_products:
            ws.cell(row=n + 1, column=4).value = len(default_products)
            for default in default_products:
                product_name, product_code = default
                ws.cell(row=n + 1, column=5).value = product_name
                ws.cell(row=n + 1, column=6).value = product_code
                n = n + 1
            merge_end = n
            ws.merge_cells(start_row=merge_start, start_column=1, end_row=merge_end, end_column=1)
            ws.merge_cells(start_row=merge_start, start_column=2, end_row=merge_end, end_column=2)
            ws.merge_cells(start_row=merge_start, start_column=3, end_row=merge_end, end_column=3)
            ws.merge_cells(start_row=merge_start, start_column=4, end_row=merge_end, end_column=4)
        else:
            n = n + 1

    wb.save('ordertypes.xlsx')

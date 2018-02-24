# coding=utf-8
"""
Created on 2018-02-07

@Filename: get_products_with_costtypes
@Author: Gui


"""
from collections import OrderedDict
from openpyxl import Workbook
from epMongo import EPMongo


def generate_pipeline():
    pipeline = [
        {
            "$match": {
                "dateDelete": {"$exists": False},
            }
        },

        {
            "$unwind": {
                "path": "$costTypes",
                "preserveNullAndEmptyArrays": True
            }
        },

        {
            "$lookup": {
                "from": "costtypes",
                "localField": "costTypes.costType",
                "foreignField": "_id",
                "as": "costTypes.costType"
            }
        },

        {
            "$unwind": {
                "path": "$costTypes.costType",
                "preserveNullAndEmptyArrays": True
            }
        },

        {
            "$project": {
                "_id": 1,
                "code": 1,
                "name": 1,
                "index": 1,
                "tags": 1,
                "parent": 1,
                "costTypeId": "$costTypes.costType._id",
                "costTypeName": "$costTypes.costType.name",
                "costTypeCode": "$costTypes.costType.code",
                "defaultVisiable": "$costTypes.defaultVisiable",
                "costTypeIndex": "$costTypes.index",
                "isEditable": "$costTypes.isEditable",
                "costTypeTags": "$costTypes.tags",

            }
        },

        {
            "$group": {
                "_id": {"_id": "$_id", "code": "$code", "name": "$name", "index": "$index", "tags": "$tags",
                        "parent": "$parent"},
                "costTypes": {
                    "$addToSet": {
                        "costTypeId": "$costTypeId",
                        "costTypeName": "$costTypeName",
                        "costTypeCode": "$costTypeCode",
                        "defaultVisiable": "$defaultVisiable",
                        "costTypeIndex": "$costTypeIndex",
                        "isEditable": "$isEditable",
                        "costTypeTags": "$costTypeTags",
                    }
                },
                "total": {"$sum": 1}
            }
        },

        {
            "$sort": {
                "_id.parent": 1,
                "_id.index": 1,
            }
        },

    ]
    return pipeline


def get_results():
    # epmongo = EPMongo()
    epmongo = EPMongo(
        uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
        db_name='epdb-prod')
    pipeline = generate_pipeline()
    results = epmongo.aggregate('products', pipeline)
    return results


def get_order_types():
    # epmongo = EPMongo()
    epmongo = EPMongo(
        uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
        db_name='epdb-prod')
    docs = epmongo.get_docs_by_query('ordertypes', {'dateDelete': {'$exists': False}}, {'name': 1, 'code': 1})
    order_types = {}
    for doc in docs:
        code = doc.get('code')
        name = doc.get('name')
        order_types[code] = name
    return order_types


def main():
    results = get_results()
    order_types = get_order_types()
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = 'No.'
    ws.cell(row=1, column=2).value = '_id'
    ws.cell(row=1, column=3).value = 'code'
    ws.cell(row=1, column=4).value = 'name'
    ws.cell(row=1, column=5).value = 'parent'
    ws.cell(row=1, column=6).value = 'index'
    ws.cell(row=1, column=7).value = 'tags'
    ws.cell(row=1, column=8).value = 'total'
    ws.cell(row=1, column=9).value = 'costTypeId'
    ws.cell(row=1, column=10).value = 'costTypeName'
    ws.cell(row=1, column=11).value = 'costTypeCode'
    ws.cell(row=1, column=12).value = 'costTypeIndex'
    ws.cell(row=1, column=13).value = 'defaultVisiable'
    ws.cell(row=1, column=14).value = 'isEditable'
    ws.cell(row=1, column=15).value = 'costTypeTags:isVisible'
    n = 1
    for i, doc in enumerate(results, 1):
        merge_start = n + 1
        ws.cell(row=n + 1, column=1).value = i
        ws.cell(row=n + 1, column=2).value = str(doc.get('_id').get('_id'))
        ws.cell(row=n + 1, column=3).value = doc.get('_id').get('code')
        ws.cell(row=n + 1, column=4).value = doc.get('_id').get('name')
        ws.cell(row=n + 1, column=5).value = str(doc.get('_id').get('parent', ''))
        ws.cell(row=n + 1, column=6).value = doc.get('_id').get('index')
        tags = doc.get('_id').get('tags')
        tags = ('\n'.join([tag.get('value', '') for tag in tags])).strip()
        ws.cell(row=n + 1, column=7).value = tags
        ws.cell(row=n + 1, column=8).value = doc.get('total')

        cost_types = doc.get('costTypes')
        if cost_types != [{}]:
            cts = sorted(cost_types, key=lambda x: x['costTypeIndex'])
            for ct in cts:
                ws.cell(row=n + 1, column=9).value = str(ct.get('costTypeId'))
                ws.cell(row=n + 1, column=10).value = ct.get('costTypeName')
                ws.cell(row=n + 1, column=11).value = ct.get('costTypeCode')
                ws.cell(row=n + 1, column=12).value = ct.get('costTypeIndex')
                ws.cell(row=n + 1, column=13).value = ct.get('defaultVisiable')
                ws.cell(row=n + 1, column=14).value = ct.get('isEditable')
                tags = ct.get('costTypeTags')
                if tags:
                    tags = ['{}:{}'.format(order_types.get(tag.get('orderTypeCode', '')), tag.get('isVisible', '')) for tag in tags if
                            order_types.get(tag.get('orderTypeCode', ''))]
                    ws.cell(row=n + 1, column=15).value = ('\n'.join(tags)).strip()
                n = n + 1
            merge_end = n
            ws.merge_cells(start_row=merge_start, start_column=1, end_row=merge_end, end_column=1)
            ws.merge_cells(start_row=merge_start, start_column=2, end_row=merge_end, end_column=2)
            ws.merge_cells(start_row=merge_start, start_column=3, end_row=merge_end, end_column=3)
            ws.merge_cells(start_row=merge_start, start_column=4, end_row=merge_end, end_column=4)
            ws.merge_cells(start_row=merge_start, start_column=5, end_row=merge_end, end_column=5)
            ws.merge_cells(start_row=merge_start, start_column=6, end_row=merge_end, end_column=6)
            ws.merge_cells(start_row=merge_start, start_column=7, end_row=merge_end, end_column=7)
            ws.merge_cells(start_row=merge_start, start_column=8, end_row=merge_end, end_column=8)
        else:
            n = n + 1
    wb.save('products.xlsx')


if __name__ == '__main__':
    main()

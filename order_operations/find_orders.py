# coding=utf-8
"""
Created on 2017-12-18

@Filename: find_orders
@Author: Gui


"""
import json
import os
import time
import logging.config
import requests
from openpyxl import Workbook
import timeit


@timeit.timeit
def get_account(env='dev', n=0):
    """
    parse accounts.yml, get url, username, password.
    :return: url, username, password
    """
    from yaml import load
    try:
        from yaml import CLoader as Loader
    except ImportError:
        from yaml import Loader
    with open('accounts.yml') as f:
        config = f.read()
    config = load(config, Loader)
    if env == 'dev':
        url = config['dev']['url']
        accounts = config['dev']['accounts']
        username = accounts[n]['username']
        password = accounts[n]['password']
    elif env == 'portal':
        url = config['portal']['url']
        accounts = config['portal']['accounts']
        username = accounts[n]['username']
        password = accounts[n]['password']
    else:
        logger.error('No {} config!'.format(env))
        raise Exception('No {} config!'.format(env))
    return url, username, password


@timeit.timeit
def get_cookie(url, username, password):
    """
    get login user cookie and role.
    :param url:
    :param username:
    :param password:
    :return: cookie, role
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
        'Content-Type': 'application/json'}
    payload = {
        'action': 'login',
        'request': {
            'username': username,
            'password': password
        }

    }
    res = requests.post(url, data=json.dumps(payload), headers=headers)
    cookie = res.headers['x-session-id']
    role = res.json()['response']['position']['group']['account']['types'][0]
    return cookie, role


@timeit.timeit
def get_orders(url, cookie):
    """
    get login user orders.
    :param url:
    :param cookie:
    :return: total, orders[(_id, orderNumber)]
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
        'Content-Type': 'application/json',
        'x-session-id': cookie
    }
    payload = {
        "action": "findOrders",
        "request": {
            "isInquiry": 'false',
            "page": "1",
            "pageSize": "4000000"
        }
    }
    res = requests.post(url, data=json.dumps(payload), headers=headers)
    orders = []
    for order in res.json()['response']['entries']:
        orders.append((order['_id'], order['orderNumber']))
    total = res.json()['response']['pagination']['total']
    return total, orders


@timeit.timeit
def find_order_by_id(url, order_id, cookie):
    """
    get order consigner and consignee.
    :param url:
    :param order_id:
    :param cookie:
    :return: consigner(_id, name), consignee(_id, name)
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
        'Content-Type': 'application/json',
        'x-session-id': cookie
    }
    payload = {
        "action": "findOrderById",
        "request": {
            "id": order_id
        }
    }
    res = requests.post(url, data=json.dumps(payload), headers=headers)
    consigner = (res.json()['response']['consigner']['_id'], res.json()['response']['consigner']['name'])
    consignee = (res.json()['response']['consignee']['_id'], res.json()['response']['consignee']['name'])
    return consigner, consignee


if __name__ == '__main__':
    if os.path.exists('mylog.log'):
        os.remove('mylog.log')
    logging.config.fileConfig('logging.conf', defaults={'logfilename': 'mylog.log'})
    logger = logging.getLogger('sLogger')
    print('main', logger)
    logger.info('Start at {}'.format(time.asctime()))
    url, username, password = get_account('dev', 0)
    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = username
    logger.info((url, username, password))
    ws['A1'] = username
    cookie, role = get_cookie(url, username, password)
    ws['B1'] = role
    total, orders = get_orders(url, cookie)
    ws['C1'] = total
    logger.info('user[{username}] role[{role}], total {total} orders'.format(username=username, role=role, total=total))
    ws['A2'] = 'orderId'
    ws['B2'] = 'orderNumber'
    ws['C2'] = 'consignerId'
    ws['D2'] = 'consignerName'
    ws['E2'] = 'consigneeId'
    ws['F2'] = 'consigneeName'
    for i, order in enumerate(orders, 1):
        logger.info('No.{i} orderNumber={orderNumber}'.format(i=i, orderNumber=order[1]))
        ws.cell(row=i + 2, column=1).value = order[0]
        ws.cell(row=i + 2, column=2).value = order[1]
        consigner, consignee = find_order_by_id(url, order[0], cookie)
        logger.info(
            'consigner --> {consigner}, consignee --> {consignee}'.format(consigner=consigner, consignee=consignee))
        ws.cell(row=i + 2, column=3).value = consigner[0]
        ws.cell(row=i + 2, column=4).value = consigner[1]
        ws.cell(row=i + 2, column=5).value = consignee[0]
        ws.cell(row=i + 2, column=6).value = consignee[1]
    wb.save('orders.xlsx')
    logger.info('End at {}'.format(time.asctime()))

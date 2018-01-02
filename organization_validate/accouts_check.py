# coding=utf-8
"""
Created on 2017-12-29

@Filename: accouts_check
@Author: Gui


"""
import logging
from bson.objectid import ObjectId
from openpyxl import load_workbook
from epMongo import EPMongo

if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG)
    # epmongo = EPMongo()
    epmongo = EPMongo(
        uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
        db_name='epdb-prod')
    file_name = '230待审核.xlsx'
    wb = load_workbook(file_name)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        organization_id = ws.cell(row=i, column=2).value
        logging.info((i-1, organization_id))
        account = epmongo.get_doc_by_query('accounts', {'organization': ObjectId(organization_id)})
        if account is None:
            ws.cell(row=i, column=5).value = 'invalid'
        else:
            account_id = account['_id']
            ws.cell(row=i, column=6).value = str(account_id)
            ws.cell(row=i, column=7).value = account['types'][0]
            username_doc = epmongo.get_doc_by_query('users', {'account': account_id, 'isSuperAdmin': True})
            if username_doc is None:
                ws.cell(row=i, column=5).value = 'redundant'
            else:
                username = username_doc['username']
                ws.cell(row=i, column=5).value = username
    wb.save(file_name)

# coding=utf-8
"""
Created on 2017-12-28

@Filename: get_organizations
@Author: Gui


"""
import logging
from openpyxl import Workbook
import pymongo
from epMongo import EPMongo

VERIFY_STATUS = {0: "待审核", 1: "审核通过", 2: "审核不通过"}


def get_organizations(verify_status=0):
    # epmongo = EPMongo()
    epmongo = EPMongo(
        uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
        db_name='epdb-prod')
    query = {"dateDelete": {"$exists": False}, "verifyStatus": verify_status}
    projection = {"_id": 1, "name": 1, "role": 1, "ports": 1}
    organizations = epmongo.get_docs_by_query("organizations", query, projection)
    # sort by dataUpdate
    organizations = organizations.sort([('dateUpdate', pymongo.DESCENDING)])
    logging.debug((organizations.count(), type(organizations)))

    ports = epmongo.get_docs_by_query("ports", None, {'name': 1, 'code': 1, '_id': 1})
    ports = {doc.get('_id'): (doc.get('code'), doc.get('name')) for doc in ports}

    organizations_list = []
    for doc in organizations:
        service_ports = doc.get('ports', '')
        if service_ports:
            str_ports = '\n'.join(
                [(ports.get(service_port)[0] + ': ' + ports.get(service_port)[1]) for service_port in service_ports])
            doc['ports'] = str_ports
        else:
            doc['ports'] = ''
        organizations_list.append(doc)
    return len(organizations_list), organizations_list


def write_to_excel(verify_status, total, organizations):
    wb = Workbook()
    ws = wb.active
    ws.title = ''.join((str(total), VERIFY_STATUS[verify_status]))
    ws.cell(row=1, column=1).value = 'No.'
    ws.cell(row=1, column=2).value = '_id'
    ws.cell(row=1, column=3).value = 'name'
    ws.cell(row=1, column=4).value = 'role'
    ws.cell(row=1, column=5).value = 'ports'
    for i, organization in enumerate(organizations, 1):
        logging.debug((i, organization))
        ws.cell(row=i + 1, column=1).value = i
        ws.cell(row=i + 1, column=2).value = str(organization['_id'])
        ws.cell(row=i + 1, column=3).value = organization['name']
        if 'role' in organization.keys():
            ws.cell(row=i + 1, column=4).value = organization['role']
        if 'ports' in organization.keys():
            ws.cell(row=i + 1, column=5).value = organization['ports']
    wb.save('%s.xlsx' % ws.title)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    verify_status = 1
    logging.info('--> get [{}] organizations'.format(VERIFY_STATUS[verify_status]))
    total, organizations = get_organizations(verify_status)
    logging.info('--> write to excel')
    write_to_excel(verify_status, total, organizations)

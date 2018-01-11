# coding=utf-8
"""
Created on 2018-01-05

@Filename: get_ships
@Author: Gui


"""
import logging
import pymongo
from openpyxl import Workbook
from epMongo import EPMongo

if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG,
                        format='%(asctime)s %(levelname)s %(message)s')
    # epmongo = EPMongo()
    epmongo = EPMongo(
        uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
        db_name='epdb-prod')
    ships = epmongo.get_docs_by_query('ships', {'dateDelete': {'$exists': False}},
                                      {'name': 1, 'imo': 1, 'type': 1, 'dateCreate': 1, 'dateUpdate': 1})
    ships = ships.sort([('imo', pymongo.ASCENDING)])
    logging.info('total {} ships.'.format(ships.count()))

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = 'No.'
    ws.cell(row=1, column=2).value = 'id'
    ws.cell(row=1, column=3).value = 'name'
    ws.cell(row=1, column=4).value = 'imo'
    ws.cell(row=1, column=5).value = 'type'
    ws.cell(row=1, column=6).value = 'dateCreate'
    ws.cell(row=1, column=7).value = 'dateUpdate'

    imos = []
    logging.info('start to generate imos list.')
    for i, ship in enumerate(ships, 1):
        ship_imo = ship.get('imo')
        imos.append(ship_imo)
    logging.info('finish to generate imos list.')
    logging.info(len(imos))
    imos_set = set(imos)
    logging.info(len(imos_set))
    i = 2
    for n, imo in enumerate(imos_set, 1):
        if imos.count(imo) != 1:
            logging.info((n, imo))
            logging.info((imo, imos.count(imo)))
            docs = epmongo.get_docs_by_query('ships', {'imo': imo, 'dateDelete': {'$exists': False}},
                                             {'name': 1, 'imo': 1, 'type': 1, 'dateCreate': 1, 'dateUpdate': 1})
            for j, ship in enumerate(docs, 1):
                logging.info((j, ship))
                ws.cell(row=i, column=1).value = i - 1
                ws.cell(row=i, column=2).value = str(ship.get('_id', '--'))
                ws.cell(row=i, column=3).value = ship.get('name', '--')
                ws.cell(row=i, column=4).value = ship.get('imo', '--')
                ws.cell(row=i, column=5).value = ship.get('type', '--')
                ws.cell(row=i, column=6).value = ship.get('dateCreate', '--')
                ws.cell(row=i, column=7).value = ship.get('dateUpdate', '--')
                i = i + 1

    wb.save('redundant_ships.xlsx')

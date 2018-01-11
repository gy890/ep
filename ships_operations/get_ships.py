# coding=utf-8
"""
Created on 2018-01-05

@Filename: get_ships
@Author: Gui


"""
import logging
import pymongo
from openpyxl import Workbook
from epMongo import EPMongo


def check_imo(imo):
    if imo == '--' or len(imo) != 7:
        return 'invalid'
    else:
        number_list = []
        for a in imo:
            number_list.append(int(a))
        result = 0
        n = len(number_list)
        for o in range(0, n - 1):
            result = result + number_list[o] * (n - o)
        if (result % 10) == number_list[-1]:
            return 'valid'
        else:
            return 'invalid'


if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s %(message)s')
    # epmongo = EPMongo()
    epmongo = EPMongo(
        uri="mongodb://root:gEUMooTG0d%23pd1%24YX@172.16.100.1:30001,172.16.100.11:30001,172.16.120.30:30001",
        db_name='epdb-prod')
    ships = epmongo.get_docs_by_query('ships', {'dateDelete': {'$exists': False}},
                                      {'name': 1, 'imo': 1, 'type': 1, 'dateCreate': 1, 'dateUpdate': 1})
    ships = ships.sort([('imo', pymongo.ASCENDING)])
    logging.info('total {} ships.'.format(ships.count()))
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = 'No.'
    ws.cell(row=1, column=2).value = 'id'
    ws.cell(row=1, column=3).value = 'name'
    ws.cell(row=1, column=4).value = 'imo'
    ws.cell(row=1, column=5).value = 'type'
    ws.cell(row=1, column=6).value = 'dateCreate'
    ws.cell(row=1, column=7).value = 'dateUpdate'
    ws.cell(row=1, column=8).value = 'imo_check'

    for i, ship in enumerate(ships, 2):
        logging.info((i - 1, ship))
        ws.cell(row=i, column=1).value = i - 1
        ws.cell(row=i, column=2).value = str(ship.get('_id', '--'))
        ws.cell(row=i, column=3).value = ship.get('name', '--')
        ws.cell(row=i, column=4).value = ship.get('imo', '--')
        ws.cell(row=i, column=5).value = ship.get('type', '--')
        ws.cell(row=i, column=6).value = ship.get('dateCreate', '--')
        ws.cell(row=i, column=7).value = ship.get('dateUpdate', '--')
        logging.debug(check_imo(ship.get('imo', '--')))
        ws.cell(row=i, column=8).value = check_imo(ship.get('imo', '--'))

    wb.save('{}.xlsx'.format(ships.count()))
